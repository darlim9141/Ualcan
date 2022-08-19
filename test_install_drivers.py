# selenium 3
#Harim Jin 2022 08 18
from ast import expr
from selenium import webdriver
from webdriver_manager.firefox import GeckoDriverManager
#from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl as op
import os

driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
#driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())

#set directory, storage file, and starting point
path = "C:\\Users\\harim\\sele\\"
download_path = "C:\\Users\\harim\\Downloads\\"
file = "store.xlsx"
row_number = 5615
last_row = 9554

#set a file for store the result
wb = op.load_workbook(path + file)
ws = wb.active

print('""""""""""""""""""""""""""""""""""""')
#row number determines the gene to import


while row_number < last_row:
    #get a gene name from file
    gene = ws.cell(row=row_number, column=1).value
    print(gene)
    #directly access to 'ualcan' web (using gene name)
    driver.get(
        f"http://ualcan.path.uab.edu/cgi-bin/TCGAExResultNew2.pl?genenam={gene}&ctype=LIHC")


    #if the gene name is not updated and can't be retrieved the web, go to 'genecards' and get previous gene name
    if not driver.find_elements(By.XPATH, "/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]"):
        driver.get(
            f"https://www.genecards.org/cgi-bin/carddisp.pl?gene={gene}&keywords={gene}")

        #check for presence of previous name of gene
        if driver.find_elements(By.XPATH, '/html/body/div[1]/div[3]/div/div/main/div[2]/div/div/section[1]/div[1]/div[1]/div[3]/div/ul'):
            #store previous gene names
            previous_genes = driver.find_element(
                By.XPATH, '/html/body/div[1]/div[3]/div/div/main/div[2]/div/div/section[1]/div[1]/div[1]/div[3]/div/ul').text.split(" ")
        else:
            row_number += 1
            print('\n""""""""""""""""""""""""""""""""""""')
            continue

        #if web with the previous gene name exist, proceed the program. if not, move on
        for i in previous_genes:
            driver.get(
                f"http://ualcan.path.uab.edu/cgi-bin/TCGAExResultNew2.pl?genenam={i}&ctype=LIHC")
            if driver.find_elements(By.XPATH, "/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]"):
                gene = i
                print(f'<<{gene}>>')
                breaker = False
                break
            breaker = True
        if breaker:
            row_number += 1
            print('\n""""""""""""""""""""""""""""""""""""')
            continue


    #search the expression value, if it is "N/A", save as string, else save as a float
    expression_value = driver.find_element(
        By.XPATH, "/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]").text

    if expression_value != "N/A" and expression_value[0] != "<":
        expression_value = float(expression_value)

    #some values includes relational operator. If so, save only the remaining 
    elif expression_value[0] == "<":
        expression_value = float(expression_value[1:])
    
    ws.cell(row=row_number, column=2).value = expression_value
    print(expression_value)


    #compare the values of normal and cancer to indicate genes with unwanted values separately
    driver.find_element(
        By.XPATH, '//*[@class="highcharts-button-symbol"]').click()  
    WebDriverWait(driver, 5).until(EC.presence_of_all_elements_located(
        (By.XPATH, '/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/div[1]/div/div/ul/li[9]')))
    driver.find_element(
        By.XPATH, '/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/div[1]/div/div/ul/li[9]').click()
    driver.implicitly_wait(1)
    nomal_low = driver.find_element(
        By.XPATH, '/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/div[2]/table/tbody/tr[1]/td[2]').text
    nomal_high = driver.find_element(
        By.XPATH, '/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/div[2]/table/tbody/tr[1]/td[4]').text
    tumor_low = driver.find_element(
        By.XPATH, '/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/div[2]/table/tbody/tr[2]/td[2]').text
    tumor_high = driver.find_element(
        By.XPATH, '/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[2]/td/div[2]/table/tbody/tr[2]/td[4]').text
    
    #if nomal lower quartile and upper quartile is greater than tumor, it is unwanting gene
    if float(nomal_low) > float(tumor_low) and float(nomal_high) >= float(tumor_high):
        ws.cell(row=row_number, column=5).value = "unwanting"
        print("unwanting")


    #search for symbols that appear when the TPM value is less than 1, save if present
    if driver.find_elements(By.XPATH, "//div[@class='tooltip']"):
        tpm_notation = driver.find_element(
            By.XPATH, "//div[@class='tooltip']").text
        ws.cell(row=row_number, column=4).value = tpm_notation
        print(tpm_notation)


    #when access the web with updated gene name, sometimes the web automatically change it with the previous gene name
    #the survival page is not automatically connected like the expression page, so use the old gene name in expression page and apply it to the survival page
    check_gene = driver.find_element(
        By.XPATH, "/html/body/div[2]/div/div[2]/div/section[2]/table/tbody/tr[1]/td/font/b").text
    check_gene = check_gene[:-22]

    #check gene name between the gene(variable) and on the web
    if check_gene != gene:
        driver.get(
            f"http://ualcan.path.uab.edu/cgi-bin/TCGA-survival1.pl?genenam={check_gene}&ctype=LIHC")
        print("<<%s>>" % check_gene)
        gene = check_gene
    else:
        driver.get(
            f"http://ualcan.path.uab.edu/cgi-bin/TCGA-survival1.pl?genenam={gene}&ctype=LIHC")
    driver.implicitly_wait(1)


    #download PDF file from the survival page if there is survival data
    if driver.find_elements(By.XPATH, "/html/body/div[2]/div/div[2]/div/section[2]/div[1]/div[1]/form/p/input[2]"):
        driver.find_element(
            By.XPATH, "/html/body/div[2]/div/div[2]/div/section[2]/div[1]/div[1]/form/p/input[2]").click()

        #close PDF file window tab
        driver.switch_to.window(driver.window_handles[1])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        #upload PDF file after accessing OCR program web page
        upload = download_path + f"{gene}-KM-Exp.pdf"
        driver.get("https://www.newocr.com")
        select = driver.find_element(
            By.XPATH, '//*[@id="userfile"]').send_keys(upload)
            

        driver.implicitly_wait(1)

        driver.find_element(By.XPATH, '//*[@id="preview"]').click()
        #wait untill the conversion is complete
        WebDriverWait(driver, 100).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[2]/div/div[2]/div[2]/div[4]/div/pre')))

        #extract only the part where the p value comes out the data returned by the web, and save it as float
        output = driver.find_element(
            By.XPATH, '/html/body/div[2]/div/div[2]/div[2]/div[4]/div/pre').text.split("\n")
        output = [i for i in output if "p " in i]
        result = "".join(output).strip()
        result = float(result[4:].replace("−", "-"))
        ws.cell(row=row_number, column=3).value = result
        print(result)

        #delete PDF file
        os.remove(download_path + "%s-KM-Exp.pdf" % gene)

        print('\n""""""""""""""""""""""""""""""""""""')
    else:
        print('\n""""""""""""""""""""""""""""""""""""')

    # save xlsx file
    wb.save(file)
    row_number = row_number + 1

driver.quit()
